import openpyxl

# 讀取excel sheet
xlsx = openpyxl.load_workbook('./menu.xlsx')
sheet = xlsx['menu']
# 第一題
beef = sheet.cell(column=1, row=2).value
pork = sheet.cell(column=2, row=2).value
lamb = sheet.cell(column=3, row=2).value

# menu dict
menu = {
    'beef': beef,
    'pork': pork,
    'lamb': lamb
}


# 更新菜單
def menu_update(i):
    global beef, pork, lamb
    if i == menu['beef'] + 1:
        beef += 1
        sheet['A2'] = beef
        print(beef)
        xlsx.save('./menu.xlsx')
    elif i == menu['beef'] - 1:
        beef -= 1
        sheet['A2'] = beef
        print(beef)
        xlsx.save('./menu.xlsx')


menu_update(menu['beef'] + 1)
