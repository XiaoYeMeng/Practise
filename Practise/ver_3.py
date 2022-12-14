import tkinter as tk
import openpyxl
import pandas as pd

window = tk.Tk()
window.title('QRCode')
window.geometry('1000x600')

# 讀取excel sheet
xlsx = openpyxl.load_workbook('./menu.xlsx')
sheet = xlsx['menu']
# 第一題
beef = sheet.cell(column=1, row=2).value
pork = sheet.cell(column=2, row=2).value
lamb = sheet.cell(column=3, row=2).value

# menu dict
menu = {
    'beef': beef,
    'pork': pork,
    'lamb': lamb
}


# 更新菜單
def menu_update(i):
    global menu
    # 如果是 XX+1 品項庫存+1
    if i == 'yuan_qi_beef_+1' and menu['beef'] >= 0:
        menu['beef'] += 1
        sheet['A2'] = menu['beef']
        xlsx.save('./menu.xlsx')
        print(menu['beef'])
    # 如果是 XX-1 品項庫存-1
    elif i == 'yuan_qi_beef_-1' and menu['beef'] > 0:
        menu['beef'] -= 1
        sheet['A2'] = menu['beef']
        xlsx.save('./menu.xlsx')
        print(menu['beef'])

    elif i == 'yuan_qi_pork_+1' and menu['pork'] >= 0:
        menu['pork'] += 1
        sheet['B2'] = menu['pork']
        xlsx.save('./menu.xlsx')
        print(menu['pork'])

    elif i == 'yuan_qi_pork_-1' and menu['pork'] > 0:
        menu['pork'] -= 1
        sheet['B2'] = menu['pork']
        xlsx.save('./menu.xlsx')
        print(menu['pork'])

    elif i == 'yuan_qi_lamb_+1' and menu['lamb'] >= 0:
        menu['lamb'] += 1
        sheet['C2'] = menu['lamb']
        xlsx.save('./menu.xlsx')
        print(menu['lamb'])

    elif i == 'yuan_qi_lamb_-1' and menu['lamb'] > 0:
        menu['lamb'] -= 1
        sheet['C2'] = menu['lamb']
        xlsx.save('./menu.xlsx')
        print(menu['lamb'])

    else:
        print('False')
    custom = pd.read_excel('menu.xlsx', sheet_name=['menu'])
    print(custom)


et = tk.Entry(window, font=('Time new roman', 18))
et.pack()
# 執行 menu_update() , 引數為 Entry 的值
bt = tk.Button(window, text='click', command=lambda: menu_update(et.get()), font=('Time new roman', 18))
bt.pack()

window.mainloop()
