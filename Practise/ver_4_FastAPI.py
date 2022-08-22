from fastapi import FastAPI
import openpyxl
from starlette.responses import FileResponse

app = FastAPI()

# 讀取excel sheet
xlsx = openpyxl.load_workbook('./menu/menu.xlsx')
sheet = xlsx['menu']
# 第一題
beef = sheet.cell(column=1, row=2).value
pork = sheet.cell(column=2, row=2).value
lamb = sheet.cell(column=3, row=2).value

# menu dict
menu = {
    'beef': beef,
    'pork': pork,
    'lamb': lamb
}


@app.post('/')
# 更新菜單
async def root(i: str):
    global menu
    # 如果是 XX+1 品項庫存+1
    if i == 'yuan_qi_beef_+1' and menu['beef'] >= 0:
        menu['beef'] += 1
        sheet['A2'] = menu['beef']
        xlsx.save('./menu/menu.xlsx')

    # 如果是 XX-1 品項庫存-1
    elif i == 'yuan_qi_beef_-1' and menu['beef'] > 0:
        menu['beef'] -= 1
        sheet['A2'] = menu['beef']
        xlsx.save('./menu/menu.xlsx')

    elif i == 'yuan_qi_pork_+1' and menu['pork'] >= 0:
        menu['pork'] += 1
        sheet['B2'] = menu['pork']
        xlsx.save('./menu/menu.xlsx')

    elif i == 'yuan_qi_pork_-1' and menu['pork'] > 0:
        menu['pork'] -= 1
        sheet['B2'] = menu['pork']
        xlsx.save('./menu/menu.xlsx')

    elif i == 'yuan_qi_lamb_+1' and menu['lamb'] >= 0:
        menu['lamb'] += 1
        sheet['C2'] = menu['lamb']
        xlsx.save('./menu/menu.xlsx')

    elif i == 'yuan_qi_lamb_-1' and menu['lamb'] > 0:
        menu['lamb'] -= 1
        sheet['C2'] = menu['lamb']
        xlsx.save('./menu/menu.xlsx')


@app.get('/image1/')
async def image():
    filename = './images/1.png'
    return FileResponse(filename, filename='fastapi_1.png')


@app.get('/image2/')
async def image():
    filename = './images/2.png'
    return FileResponse(filename, filename='fastapi_2.png')


@app.get('/image3/')
async def image():
    filename = './images/3.png'
    return FileResponse(filename, filename='fastapi_3.png')
